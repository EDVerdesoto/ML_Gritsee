from typing import List, Optional
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends, Query
from sqlalchemy.orm import Session
from datetime import date, datetime
import pandas as pd
import io
from fastapi.responses import StreamingResponse
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from app.db.session import get_db
from app.models.inspeccion import Inspeccion
from app.schemas.inspeccion_schema import InspeccionResponse, InspeccionUpdate
from app.services.quality_service import QualityService
from app.services.scoring_logic import calcular_puntaje
from app.services.inspeccion_service import InspeccionService

router = APIRouter()

# ==========================================
# 1. CARGA MASIVA (BATCH UPLOAD)
# ==========================================
@router.post("/batch-upload")
async def cargar_csv_inspecciones(
    file: UploadFile = File(...),
    locacion: str = Form(...), # <--- AQUÍ RECIBIMOS LA LOCACIÓN DEL FRONT
    db: Session = Depends(get_db)
):
    # 1. Validación
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo archivos CSV o Excel")

    # 2. Lectura
    contents = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df_raw = pd.read_csv(io.BytesIO(contents), header=None)  
        else:
            df_raw = pd.read_excel(io.BytesIO(contents), header=None)
    except Exception as e:
        print(f"Error lectura: {e}")
        raise HTTPException(status_code=400, detail="Archivo corrupto")

    # 3. Header Discovery (Buscar dónde empieza)
    header_row_index = -1
    for i, row in df_raw.head(20).iterrows():
        row_str = row.astype(str).str.lower().tolist()
        if any("link" in s for s in row_str): # Buscamos "Photo Link" o similar
            header_row_index = i
            break
    
    if header_row_index == -1:
        raise HTTPException(status_code=400, detail="No encontré cabeceras (Photo Link)")

    # 4. Limpieza y Asignación de columnas
    df_raw.columns = df_raw.iloc[header_row_index]
    df = df_raw.iloc[header_row_index + 1 :].reset_index(drop=True)
    
    # Identificar columna Link y Columna Fecha
    col_link = next((c for c in df.columns if "link" in str(c).lower()), None)
    col_fecha = next((c for c in df.columns if "fecha" in str(c).lower() or "date" in str(c).lower()), None)
    # CHECK 5: Identificar columna de Hora separada
    col_hora = next((c for c in df.columns if "hora" in str(c).lower() or "time" in str(c).lower()), None)

    if not col_link:
        raise HTTPException(status_code=400, detail="No encontré columna de Links")

    # 5. PREPARAR DATOS PARA EL SERVICIO
    # En lugar de mandar solo links, mandamos una lista de diccionarios con fecha
    datos_procesar = []
    
    for _, row in df.iterrows():
        link = str(row[col_link]).strip()
        if not link.startswith("http"): continue
        
        # Intentar parsear fecha, si no existe o falla, usa HOY
        fecha_obj = datetime.now()
        if col_fecha and pd.notna(row[col_fecha]):
            try:
                # Pandas es inteligente parseando fechas
                fecha_obj = pd.to_datetime(row[col_fecha]).to_pydatetime()
            except:
                pass # Si falla el parseo, se queda con la fecha de hoy
        
        # CHECK 5: Concatenar hora si existe columna separada
        if col_hora and pd.notna(row[col_hora]):
            try:
                hora_str = str(row[col_hora]).strip()
                # Parsear diferentes formatos de hora: "12:30", "12:30:00", "1230"
                if ":" in hora_str:
                    partes = hora_str.split(":")
                    hora = int(partes[0])
                    minuto = int(partes[1]) if len(partes) > 1 else 0
                    segundo = int(partes[2]) if len(partes) > 2 else 0
                else:
                    # Formato sin separador (ej: "1230")
                    hora = int(hora_str[:2]) if len(hora_str) >= 2 else 0
                    minuto = int(hora_str[2:4]) if len(hora_str) >= 4 else 0
                    segundo = 0
                
                # Combinar fecha + hora
                fecha_obj = fecha_obj.replace(hour=hora, minute=minuto, second=segundo)
            except Exception as e:
                print(f"⚠️ No se pudo parsear hora '{row[col_hora]}': {e}")
                # Mantiene la fecha sin hora específica
        
        datos_procesar.append({
            "link": link,
            "fecha": fecha_obj
        })

    # 6. INVOCAR AL SERVICIO
    servicio = QualityService(db)
    # OJO: Tienes que actualizar tu quality_service para que acepte este nuevo formato
    # (Ver instrucciones abajo)
    resultados = servicio.procesar_lista_con_metadata(datos_procesar, locacion)

    return {
        "status": "OK",
        "total_procesados": len(resultados),
        "detalle": resultados
    }


# ==========================================
# 2. LISTADO Y FILTROS (GET)
# ==========================================
@router.get("/", response_model=List[InspeccionResponse])
def leer_inspecciones(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 50,
    # Filtros
    id: Optional[str] = None,
    locacion: Optional[str] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    veredicto: Optional[str] = None,
    # Ordenamiento
    sort_by: Optional[str] = "id",
    sort_order: Optional[str] = "desc"
):
    # 1. USAMOS EL SERVICIO
    query = InspeccionService.aplicar_filtros(
        db=db,
        id=id,
        locacion=locacion,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        veredicto=veredicto
    )

    # 2. APLICAMOS ORDENAMIENTO (Esto suele ser específico del endpoint)
    # Validamos que la columna exista para evitar inyecciones o errores
    if hasattr(Inspeccion, sort_by):
        order_column = getattr(Inspeccion, sort_by)
    else:
        order_column = Inspeccion.id

    if sort_order == "asc":
        query = query.order_by(order_column.asc())
    else:
        query = query.order_by(order_column.desc())
    
    # 3. PAGINACIÓN Y EJECUCIÓN
    registros = query.offset(skip).limit(limit).all()
    
    return registros


# ==========================================
# 3. DETALLE ÚNICO (GET)
# ==========================================
@router.get("/{id}", response_model=InspeccionResponse)
def leer_inspeccion_detalle(id: int, db: Session = Depends(get_db)):
    inspeccion = db.query(Inspeccion).filter(Inspeccion.id == id).first()
    if not inspeccion:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    return inspeccion


# ==========================================
# 4. CORRECCIÓN HUMANA (PATCH)
# ==========================================
@router.patch("/{id}", response_model=InspeccionResponse)
def corregir_inspeccion(
    id: int,
    datos_update: InspeccionUpdate,
    db: Session = Depends(get_db)
):
    inspeccion = db.query(Inspeccion).filter(Inspeccion.id == id).first()
    if not inspeccion:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")

    update_data = datos_update.dict(exclude_unset=True)
    
    # Convertir 0/1 a bool si vienen como enteros
    bool_fields = ['tiene_burbujas', 'bordes_sucios', 'tiene_grasa']
    for field in bool_fields:
        if field in update_data and isinstance(update_data[field], int):
            update_data[field] = bool(update_data[field])
    
    for key, value in update_data.items():
        setattr(inspeccion, key, value)

    # RE-SCORING
    datos_para_scoring = {
        "tiene_burbujas": inspeccion.tiene_burbujas,
        "bordes_sucios": inspeccion.bordes_sucios, 
        "tiene_grasa": inspeccion.tiene_grasa,
        "horneado": inspeccion.horneado_clase,
        "distribucion": inspeccion.distribucion_clase,
        "bordes_limpios": not inspeccion.bordes_sucios 
    }
    
    nuevos_scores = calcular_puntaje(datos_para_scoring)
    
    inspeccion.score_burbujas = nuevos_scores['burbujas']
    inspeccion.score_bordes = nuevos_scores['bordes']
    inspeccion.score_grasa = nuevos_scores['grasa']
    inspeccion.score_horneado = nuevos_scores['horneado']
    inspeccion.score_distribucion = nuevos_scores['distribucion']
    inspeccion.puntaje_total = nuevos_scores['total']
    inspeccion.veredicto = nuevos_scores['veredicto']
    
    db.commit()
    db.refresh(inspeccion)
    return inspeccion


# ==========================================
# 5. OPCIONES PARA FILTROS (METADATA)
# ==========================================
@router.get("/opciones/metadata")
def obtener_opciones_filtro(db: Session = Depends(get_db)):
    """
    Esto alimenta el Dropdown del Frontend.
    """
    locaciones = db.query(Inspeccion.locacion).distinct().all()
    lista_locaciones = [loc[0] for loc in locaciones if loc[0]]
    return {"locaciones": sorted(lista_locaciones)}

# ==========================================
# 6. EXPORTAR A EXCEL 
# ==========================================

@router.get("/exportar/excel")
def exportar_inspecciones_excel(
    # Recibimos los mismos filtros que usas en el GET normal
    id: Optional[str] = None,
    locacion: Optional[str] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    veredicto: Optional[str] = None,
    limit: int = 999999, # Límite alto para exportación
    db: Session = Depends(get_db),
):
    # 1. REUTILIZAMOS LA LÓGICA
    query = InspeccionService.aplicar_filtros(
        db=db,
        id=id,
        locacion=locacion,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        veredicto=veredicto
    )
    
    # 2. Ordenar y obtener datos
    resultados = query.order_by(Inspeccion.fecha_hora.desc()).limit(limit).all()
    
    # 2. Convertir a DataFrame de Pandas (Para manipular fácil)
    data = []
    for ins in resultados:
        data.append({
            "ID": ins.id,
            "Fecha": ins.fecha_hora, # Aquí podrías ajustar Timezone si quisieras
            "Sucursal": ins.locacion,
            "Puntaje": ins.puntaje_total,
            "Veredicto": "PASS" if ins.puntaje_total >= 80 else "FAIL",
            "Burbujas": "Sí" if ins.tiene_burbujas else "No",
            "Bordes Sucios": "Sí" if ins.bordes_sucios else "No",
            "Horneado": ins.horneado_clase,
            "Distribución": ins.distribucion_clase,
            "Grasa": "Sí" if ins.tiene_grasa else "No",
            "Link Imagen": ins.aws_link
        })
        
    df = pd.DataFrame(data)
    
    # 3. Generar el archivo en memoria (BytesIO)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Inspecciones')
        
        # Obtener la hoja activa para aplicar estilos
        workbook = writer.book
        worksheet = writer.sheets['Inspecciones']
        
        # Estilo de borde
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color='AFC7A5', end_color='1F2937', fill_type='solid')
        header_font = Font(bold=True, color='000000', size=11)
        
        # Aplicar estilo a encabezados
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar bordes y alineación a todas las celdas de datos
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # Auto-ajustar ancho de columnas con lógica mejorada
        for idx, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = column[0].column_letter
            column_name = column[0].value  # Nombre del encabezado
            
            for cell in column:
                try:
                    if cell.value:
                        # Calcular longitud considerando el contenido
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Lógica especial para columna "Link Imagen" (última columna)
            if column_name == "Link Imagen":
                # Aplicar wrap_text (ajuste de texto) a todas las celdas de esta columna
                for cell in column:
                    if cell.row > 1:  # Saltar encabezado
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                # Ancho fijo para links (suficiente para mostrar bien con wrap)
                worksheet.column_dimensions[column_letter].width = 40
                # Ajustar altura de filas automáticamente
                for row_idx in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_idx].height = None  # Auto
            else:
                # Para otras columnas, ajustar al contenido con límite razonable
                adjusted_width = min(max_length + 3, 40)  # Máximo 40 caracteres
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # 4. Devolver como descarga
    headers = {
        'Content-Disposition': f'attachment; filename="reporte_gritsee_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    }
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
